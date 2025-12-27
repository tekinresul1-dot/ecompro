"""
Products App - Views
"""

from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models import Q

from .models import Product, ProductCostHistory, BulkCostUpload
from .serializers import (
    ProductSerializer,
    ProductCostUpdateSerializer,
    ProductCostHistorySerializer,
    BulkCostUploadSerializer,
    BulkCostUploadCreateSerializer,
)


class ProductListView(generics.ListAPIView):
    """
    List products for the current user's seller accounts.
    
    Filters:
    - seller_account: Filter by seller account ID
    - has_cost_data: Filter by cost data availability
    - is_active: Filter by active status
    - search: Search by barcode, product code, or title
    - category: Filter by category
    - brand: Filter by brand
    """
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = Product.objects.filter(
            seller_account__user=user
        ).select_related('seller_account')
        
        # Apply filters
        params = self.request.query_params
        
        seller_account = params.get('seller_account')
        if seller_account:
            queryset = queryset.filter(seller_account_id=seller_account)
        
        has_cost_data = params.get('has_cost_data')
        if has_cost_data is not None:
            queryset = queryset.filter(has_cost_data=has_cost_data.lower() == 'true')
        
        is_active = params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(barcode__icontains=search) |
                Q(product_code__icontains=search) |
                Q(title__icontains=search)
            )
        
        category = params.get('category')
        if category:
            queryset = queryset.filter(category__icontains=category)
        
        brand = params.get('brand')
        if brand:
            queryset = queryset.filter(brand__icontains=brand)
        
        return queryset


class ProductDetailView(generics.RetrieveAPIView):
    """
    Get product details.
    """
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return Product.objects.filter(
            seller_account__user=self.request.user
        )


class ProductCostUpdateView(APIView):
    """
    Update product cost.
    """
    permission_classes = [IsAuthenticated]
    
    def patch(self, request, pk):
        try:
            product = Product.objects.get(
                pk=pk,
                seller_account__user=request.user
            )
        except Product.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Ürün bulunamadı.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        serializer = ProductCostUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        
        # Update cost with history tracking
        product.update_cost(
            cost_excl_vat=data['product_cost_excl_vat'],
            vat_rate=data.get('purchase_vat_rate'),
            track_history=True
        )
        
        # Update other fields if provided
        if 'sales_vat_rate' in data:
            product.sales_vat_rate = data['sales_vat_rate']
        if 'commission_rate' in data:
            product.commission_rate = data['commission_rate']
        product.save()
        
        return Response({
            'success': True,
            'message': 'Ürün maliyeti güncellendi.',
            'data': ProductSerializer(product).data
        })


class ProductCostHistoryView(generics.ListAPIView):
    """
    Get cost history for a product.
    """
    serializer_class = ProductCostHistorySerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        product_id = self.kwargs.get('pk')
        return ProductCostHistory.objects.filter(
            product_id=product_id,
            product__seller_account__user=self.request.user
        )


class BulkCostUploadView(APIView):
    """
    Upload Excel file for bulk cost update.
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def post(self, request):
        serializer = BulkCostUploadCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        seller_account_id = serializer.validated_data['seller_account']
        file = serializer.validated_data['file']
        
        # Verify seller account belongs to user
        from apps.sellers.models import SellerAccount
        try:
            seller_account = SellerAccount.objects.get(
                pk=seller_account_id,
                user=request.user
            )
        except SellerAccount.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Satıcı hesabı bulunamadı.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Create upload record
        upload = BulkCostUpload.objects.create(
            seller_account=seller_account,
            file_name=file.name,
            file_path=file,
            status='pending'
        )
        
        # Queue processing task
        from apps.products.tasks import process_bulk_cost_upload
        process_bulk_cost_upload.delay(upload.id)
        
        return Response({
            'success': True,
            'message': 'Dosya yüklendi, işleme alındı.',
            'data': BulkCostUploadSerializer(upload).data
        }, status=status.HTTP_202_ACCEPTED)


class BulkCostUploadStatusView(APIView):
    """
    Get status of a bulk upload.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk):
        try:
            upload = BulkCostUpload.objects.get(
                pk=pk,
                seller_account__user=request.user
            )
        except BulkCostUpload.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Yükleme kaydı bulunamadı.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        return Response({
            'success': True,
            'data': BulkCostUploadSerializer(upload).data
        })


class BulkCostUploadListView(generics.ListAPIView):
    """
    List bulk uploads for a seller account.
    """
    serializer_class = BulkCostUploadSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        seller_id = self.request.query_params.get('seller_account')
        queryset = BulkCostUpload.objects.filter(
            seller_account__user=self.request.user
        )
        if seller_id:
            queryset = queryset.filter(seller_account_id=seller_id)
        return queryset.order_by('-created_at')[:20]


class ProductsWithoutCostView(generics.ListAPIView):
    """
    List products without cost data.
    Useful for identifying products that need cost entry.
    """
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        seller_id = self.request.query_params.get('seller_account')
        queryset = Product.objects.filter(
            seller_account__user=self.request.user,
            has_cost_data=False,
            is_active=True
        )
        if seller_id:
            queryset = queryset.filter(seller_account_id=seller_id)
        return queryset


class ProductExportView(APIView):
    """
    Export products to Excel.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        seller_id = request.query_params.get('seller_account')
        queryset = Product.objects.filter(
            seller_account__user=request.user
        ).select_related('seller_account')
        
        if seller_id:
            queryset = queryset.filter(seller_account_id=seller_id)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ürünler'
        
        # Headers
        headers = [
            'Barkod', 'Model Kodu', 'Ürün Adı', 'Marka', 'Kategori', 'Görsel Linki',
            'Ürün Maliyeti (KDV Dahil)', 'Maliyet (KDV Hariç)', 'Maliyet KDV Oranı', 
            'Satış KDV Oranı', 'Komisyon Oranı', 'Mağaza'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row, product in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=product.barcode)
            ws.cell(row=row, column=2, value=product.product_code)
            ws.cell(row=row, column=3, value=product.title)
            ws.cell(row=row, column=4, value=product.brand)
            ws.cell(row=row, column=5, value=product.category)
            ws.cell(row=row, column=6, value=product.image_url)
            
            # Calculate Cost Inc VAT
            cost_excl = float(product.product_cost_excl_vat) if product.product_cost_excl_vat else 0
            vat_rate = float(product.purchase_vat_rate) if product.purchase_vat_rate else 0
            cost_inc = cost_excl * (1 + vat_rate / 100)
            
            ws.cell(row=row, column=7, value=cost_inc if cost_excl > 0 else None)
            ws.cell(row=row, column=8, value=cost_excl if cost_excl > 0 else None)
            ws.cell(row=row, column=9, value=vat_rate)
            ws.cell(row=row, column=10, value=float(product.sales_vat_rate))
            ws.cell(row=row, column=11, value=float(product.commission_rate) if product.commission_rate else None)
            ws.cell(row=row, column=12, value=product.seller_account.shop_name)
        
        # Auto-width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=urunler.xlsx'
        wb.save(response)
        
        return response

class ProductUpdateFromExcelView(APIView):
    """
    Update product details (Image, Brand) from standard Trendyol Excel export.
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def post(self, request):
        import openpyxl
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'success': False, 'message': 'Dosya yüklenmedi.'}, status=400)
            
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
            
            # Map headers
            header_map = {}
            for cell in ws[1]:
                if cell.value:
                    header_map[str(cell.value).strip()] = cell.column - 1
            
            required = ['Barkod']
            missing = [h for h in required if h not in header_map]
            if missing:
                 return Response({'success': False, 'message': f'Eksik sütunlar: {", ".join(missing)}. "Barkod" sütunu gereklidir.'}, status=400)
            
            col_barcode = header_map['Barkod']
            # Try to find Image and Brand columns
            col_image = header_map.get('Görsel Linki') or header_map.get('Görsel Linkleri') or header_map.get('Görsel 1')
            col_brand = header_map.get('Marka')
            
            updated_count = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_barcode = row[col_barcode]
                if not raw_barcode:
                    continue
                
                # Clean barcode (handle float from Excel)
                barcode = str(raw_barcode).strip()
                if barcode.endswith('.0'):
                    barcode = barcode[:-2]
                    
                updates = {}
                
                # Image
                if col_image is not None and row[col_image]:
                     img_url = str(row[col_image]).strip()
                     # If multiple images (comma separated or newline), take first
                     if ',' in img_url:
                          img_url = img_url.split(',')[0].strip()
                     if '\n' in img_url:
                          img_url = img_url.split('\n')[0].strip()
                     if img_url:
                          updates['image_url'] = img_url
                
                # Brand
                if col_brand is not None and row[col_brand]:
                     brand = str(row[col_brand]).strip()
                     if brand:
                          updates['brand'] = brand
                          
                if updates:
                     # Update matching products
                     Product.objects.filter(
                         seller_account__user=request.user, 
                         barcode=barcode
                     ).update(**updates)
                     updated_count += 1
                     
            return Response({
                'success': True, 
                'message': f'{updated_count} ürün bilgisi Excel\'den güncellendi.'
            })
            
        except Exception as e:
            # Handle specific openpyxl errors or generic
            return Response({'success': False, 'message': f'Dosya işlenirken hata: {str(e)}'}, status=400)
