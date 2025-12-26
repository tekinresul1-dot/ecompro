"""
Products App - Celery Tasks
"""

from celery import shared_task
import logging

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def process_bulk_cost_upload(self, upload_id: int):
    """
    Process a bulk cost upload Excel file.
    
    Expected Excel format:
    | Barkod | Maliyet (KDV Hariç) | Alış KDV % | Komisyon % |
    """
    from django.utils import timezone
    import openpyxl
    
    from .models import BulkCostUpload, Product
    
    try:
        upload = BulkCostUpload.objects.get(pk=upload_id)
    except BulkCostUpload.DoesNotExist:
        logger.error(f'Bulk upload {upload_id} not found')
        return
    
    upload.status = 'processing'
    upload.save(update_fields=['status'])
    
    errors = []
    success_count = 0
    
    try:
        # Load Excel file
        wb = openpyxl.load_workbook(upload.file_path.path)
        ws = wb.active
        
        # Get total rows (excluding header)
        total_rows = ws.max_row - 1
        upload.total_rows = total_rows
        upload.save(update_fields=['total_rows'])
        
        # Process each row
        for row_num in range(2, ws.max_row + 1):
            try:
                barcode = str(ws.cell(row=row_num, column=1).value or '').strip()
                cost_value = ws.cell(row=row_num, column=2).value
                vat_rate_value = ws.cell(row=row_num, column=3).value
                commission_value = ws.cell(row=row_num, column=4).value
                
                if not barcode:
                    continue
                
                # Find product
                product = Product.objects.filter(
                    seller_account=upload.seller_account,
                    barcode=barcode
                ).first()
                
                if not product:
                    errors.append({
                        'row': row_num,
                        'barcode': barcode,
                        'error': 'Ürün bulunamadı'
                    })
                    continue
                
                # Parse and validate cost
                if cost_value is None:
                    errors.append({
                        'row': row_num,
                        'barcode': barcode,
                        'error': 'Maliyet değeri boş'
                    })
                    continue
                
                try:
                    from decimal import Decimal
                    cost = Decimal(str(cost_value))
                    if cost < 0:
                        raise ValueError('Negatif değer')
                except (ValueError, TypeError):
                    errors.append({
                        'row': row_num,
                        'barcode': barcode,
                        'error': f'Geçersiz maliyet değeri: {cost_value}'
                    })
                    continue
                
                # Parse VAT rate (optional)
                vat_rate = Decimal('20.00')
                if vat_rate_value is not None:
                    try:
                        vat_rate = Decimal(str(vat_rate_value))
                    except (ValueError, TypeError):
                        pass
                
                # Parse commission rate (optional)
                if commission_value is not None:
                    try:
                        product.commission_rate = Decimal(str(commission_value))
                    except (ValueError, TypeError):
                        pass
                
                # Update product
                product.update_cost(cost, vat_rate, track_history=True)
                success_count += 1
                
                # Update progress
                upload.processed_rows = row_num - 1
                upload.save(update_fields=['processed_rows'])
                
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'barcode': barcode if 'barcode' in locals() else 'N/A',
                    'error': str(e)
                })
        
        # Mark as completed
        upload.status = 'completed'
        upload.success_count = success_count
        upload.error_count = len(errors)
        upload.error_details = errors[:100]  # Limit stored errors
        upload.processed_at = timezone.now()
        upload.save()
        
        logger.info(f'Bulk upload {upload_id} completed: {success_count} success, {len(errors)} errors')
        
    except Exception as e:
        logger.exception(f'Bulk upload {upload_id} failed: {e}')
        upload.status = 'failed'
        upload.error_details = [{'error': str(e)}]
        upload.save(update_fields=['status', 'error_details'])
        
        # Retry on transient errors
        raise self.retry(exc=e, countdown=60)
